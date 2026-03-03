from openpyxl import load_workbook
import pandas as pd
import shutil
from datetime import datetime

PRODUCT_FILE = "Product_file.xlsx"
DICT_FILE = "filter_dictionary.xlsx"

SKU_COL = 2
FILTER_START_COL = 85
FILTER_HEADER_ROW = 3
DATA_START_ROW = 4

# ---------------------------------
# Create Backup
# ---------------------------------
timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
backup_name = f"backup_{timestamp}_Product_file.xlsx"
shutil.copy(PRODUCT_FILE, backup_name)
print(f"Backup created: {backup_name}")

# ---------------------------------
# Create Log File
# ---------------------------------
log_filename = f"change_log_{timestamp}.txt"
log_file = open(log_filename, "w", encoding="utf-8")

log_file.write(f"Change Log - {datetime.now()}\n")
log_file.write(f"Product File: {PRODUCT_FILE}\n")
log_file.write(f"Dictionary File: {DICT_FILE}\n")
log_file.write("=" * 60 + "\n\n")

# ---------------------------------
# Load dictionary
# ---------------------------------
dictionary = pd.read_excel(DICT_FILE)
dictionary = dictionary.dropna(subset=["Canonical Value"])

def norm(x):
    return str(x).strip().lower()

dictionary["Filter Name"] = dictionary["Filter Name"].apply(norm)
dictionary["Filter Value"] = dictionary["Filter Value"].apply(norm)

mapping = {
    (row["Filter Name"], row["Filter Value"]): row["Canonical Value"].strip()
    for _, row in dictionary.iterrows()
}

print(f"Loaded {len(mapping)} dictionary mappings.")
log_file.write(f"Loaded {len(mapping)} dictionary mappings.\n\n")

# ---------------------------------
# Load workbook
# ---------------------------------
wb = load_workbook(PRODUCT_FILE)

total_changes = 0

for ws in wb.worksheets:
    print(f"\nProcessing sheet: {ws.title}")
    log_file.write(f"\nProcessing sheet: {ws.title}\n")

    # Read headers
    filter_headers = []
    for col in range(FILTER_START_COL, ws.max_column + 1):
        val = ws.cell(row=FILTER_HEADER_ROW, column=col).value
        filter_headers.append(norm(val) if val else "")

    sheet_changes = 0

    for row in range(DATA_START_ROW, ws.max_row + 1):
        sku = ws.cell(row=row, column=SKU_COL).value  # ← Get SKU once per row

        for offset, filter_name in enumerate(filter_headers):
            col = FILTER_START_COL + offset
            cell = ws.cell(row=row, column=col)

            if cell.value is None:
                continue

            original = str(cell.value).strip()
            key = (filter_name, original.lower())

            if key in mapping:
                new_val = mapping[key]
                if original != new_val:
                    cell.value = new_val
                    sheet_changes += 1

                    # Log the change including SKU
                    log_file.write(
                        f"Sheet: {ws.title} | "
                        f"SKU: {sku} | "
                        f"Row: {row} | "
                        f"Column: {col} | "
                        f"Filter: {filter_name} | "
                        f"Old: '{original}' → New: '{new_val}'\n"
                    )

    print(f"  → {sheet_changes} changes")
    log_file.write(f"  → {sheet_changes} changes in sheet\n")

    total_changes += sheet_changes

# ---------------------------------
# Save workbook
# ---------------------------------
wb.save(PRODUCT_FILE)

log_file.write("\n" + "=" * 60 + "\n")
log_file.write(f"Total changes across all sheets: {total_changes}\n")
log_file.write(f"Completed at: {datetime.now()}\n")
log_file.close()

print(f"\nDone. Total changes across all sheets: {total_changes}")
print(f"Log saved to: {log_filename}")